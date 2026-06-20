
from __future__ import annotations

from datetime import date, datetime, timedelta
from io import BytesIO
from pathlib import Path
from collections import defaultdict
import math
import os
import re

import pandas as pd
import xlsxwriter
from openpyxl import load_workbook


DAY_INDEX = {
    "monday": 0,
    "tuesday": 1,
    "wednesday": 2,
    "thursday": 3,
    "friday": 4,
    "saturday": 5,
    "sunday": 6,
}

SKIP_TERMS = (
    "lunch", "week off", "weekend off", "holiday", "self study", "self-study",
    "rest time", "nap", "football", "t.t", "tt time", "break", "meeting",
    "ptm", "test jee", "exam prep", "academic meeting", "solving slot",
)

TEACHER_TITLE_RE = re.compile(r"\b(?:sir|ma['’]?am|maam|mam|madam)\b", re.IGNORECASE)

TIME_RANGE_RE = re.compile(
    r"(?i)"
    r"(?P<start>\d{1,2}(?:\s*[:.]\s*\d{1,2})?\s*(?:a\.?m\.?|p\.?m\.?)?)"
    r"\s*(?:-|–|—|to)\s*"
    r"(?P<end>\d{1,2}(?:\s*[:.]\s*\d{1,2})?\s*(?:a\.?m\.?|p\.?m\.?)?)"
)

SUBJECT_PATTERNS = [
    ("Doubt - Mathematics", r"^doubt\s*[-–—]?\s*(?:math|maths|mathematics)\s*[-:–—]?\s*"),
    ("Doubt - Physics", r"^doubt\s*[-–—]?\s*(?:phy|physics)\s*[-:–—]?\s*"),
    ("Doubt - Chemistry", r"^doubt\s*[-–—]?\s*(?:chem|chemistry)\s*[-:–—]?\s*"),
    ("Doubt - Biology", r"^doubt\s*[-–—]?\s*(?:bio|biology)\s*[-:–—]?\s*"),
    ("Online - Mathematics", r"^online\s*[-–—]?\s*(?:math|maths|mathematics)\s*[-:–—]?\s*"),
    ("Online - MAT", r"^online\s*[-–—]?\s*mat\s*[-:–—]?\s*"),
    ("Online - SST", r"^online\s*[-–—]?\s*(?:sst|social science)\s*[-:–—]?\s*"),
    ("Online - English", r"^online\s*[-–—]?\s*(?:eng|english)\s*[-:–—]?\s*"),
    ("Mathematics Olympiad", r"^maths?\s+olympiad\s*"),
    ("Physics Olympiad", r"^physics\s+olympiad\s*"),
    ("Chemistry Olympiad", r"^chemistry\s+olympiad\s*"),
    ("Biology Olympiad", r"^biology\s+olympiad\s*"),
    ("Physics", r"^(?:phy|physics)\s*[-:–—]\s*"),
    ("Chemistry", r"^(?:chem|chemistry)\s*[-:–—]\s*"),
    ("Mathematics", r"^(?:math|maths|mathematics)\s*[-:–—]\s*"),
    ("Biology", r"^(?:bio|biology)\s*[-:–—]\s*"),
    ("SST", r"^(?:sst|social science)\s*[-:–—]\s*"),
    ("English", r"^(?:eng|english)\s*[-:–—]\s*"),
    ("NSEP", r"^nsep\s*[-:–—]\s*"),
    ("NSEJS", r"^nsejs\s*[-:–—]\s*"),
    ("IOQM", r"^ioqm\s*[-:–—]\s*"),
    ("RMO", r"^rmo\s*[-:–—]\s*"),
    ("NEET", r"^neet\s*[-:–—]\s*"),
    ("JEE", r"^jee\s*[-:–—]\s*"),
]


def clean_text(value) -> str:
    if value is None:
        return ""
    text = str(value).replace("\xa0", " ").replace("\n", " ").replace("\r", " ")
    return re.sub(r"\s+", " ", text).strip()


def excel_column_letter(number: int) -> str:
    result = ""
    while number:
        number, remainder = divmod(number - 1, 26)
        result = chr(65 + remainder) + result
    return result


def cell_value(matrix: list[list], row: int, col: int):
    if row < 0 or col < 0 or row >= len(matrix) or col >= len(matrix[row]):
        return ""
    return matrix[row][col]


def normalise_teacher(value: str) -> str:
    result = clean_text(value).lower()
    result = result.replace("ma'am", "maam").replace("ma’am", "maam")
    result = re.sub(r"[^a-z0-9 ]+", " ", result)
    return re.sub(r"\s+", " ", result).strip()


def is_skip_entry(value: str) -> bool:
    lower = clean_text(value).lower()
    return any(term in lower for term in SKIP_TERMS)


def looks_like_teacher_name(value: str) -> bool:
    return bool(TEACHER_TITLE_RE.search(clean_text(value)))


def parse_date_value(value):
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        if 30000 <= float(value) <= 70000:
            return (datetime(1899, 12, 30) + timedelta(days=float(value))).date()

    text = clean_text(value)
    for fmt in (
        "%d.%m.%Y", "%d-%m-%Y", "%d/%m/%Y",
        "%Y-%m-%d", "%d.%m.%y", "%d-%m-%y", "%d/%m/%y",
    ):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    return None


def get_day_name(value: str):
    lower = clean_text(value).lower()
    for day in DAY_INDEX:
        if day in lower:
            return day
    return None


def parse_time_token(token: str):
    text = clean_text(token).lower().replace(" ", "")
    suffix = None
    if "a.m." in text or "am" in text:
        suffix = "am"
    elif "p.m." in text or "pm" in text:
        suffix = "pm"

    text = re.sub(r"(a\.?m\.?|p\.?m\.?)", "", text).replace(".", ":")
    if ":" in text:
        hour_text, minute_text = text.split(":", 1)
    else:
        hour_text, minute_text = text, "0"

    try:
        hour = int(hour_text)
        minute = int(minute_text)
    except ValueError:
        return None

    if not (0 <= hour <= 24 and 0 <= minute <= 59):
        return None
    if hour == 0:
        hour = 12
    return hour, minute, suffix


def to_minutes(hour: int, minute: int, suffix: str) -> int:
    if suffix == "am":
        hour_24 = 0 if hour == 12 else hour
    else:
        hour_24 = 12 if hour == 12 else hour + 12
    return hour_24 * 60 + minute


def infer_time_range(start_token: str, end_token: str):
    start = parse_time_token(start_token)
    end = parse_time_token(end_token)
    if not start or not end:
        return None

    sh, sm, ssuffix = start
    eh, em, esuffix = end

    # Timetable convention: 8–11 are morning; 12–7 are afternoon/evening.
    if ssuffix is None and esuffix is None:
        ssuffix = "am" if 8 <= sh <= 11 else "pm"
        esuffix = "pm" if eh == 12 else ssuffix
    elif ssuffix is not None and esuffix is None:
        esuffix = "pm" if ssuffix == "am" and eh == 12 else ssuffix
    elif ssuffix is None and esuffix is not None:
        if esuffix == "am":
            ssuffix = "am"
        elif sh == 12:
            ssuffix = "pm"
        elif 8 <= sh <= 11 and eh == 12:
            ssuffix = "am"
        else:
            ssuffix = "pm"

    start_minutes = to_minutes(sh, sm, ssuffix)
    end_minutes = to_minutes(eh, em, esuffix)

    if end_minutes <= start_minutes and ssuffix == "am" and eh == 12:
        end_minutes = to_minutes(eh, em, "pm")
    if end_minutes <= start_minutes and end_minutes < 12 * 60:
        end_minutes += 12 * 60

    if not (0 <= start_minutes < 24 * 60 and 0 < end_minutes <= 24 * 60):
        return None
    if end_minutes <= start_minutes:
        return None
    return start_minutes, end_minutes


def time_range_from_text(value: str):
    matches = list(TIME_RANGE_RE.finditer(clean_text(value)))
    if not matches:
        return None
    match = matches[-1]
    return infer_time_range(match.group("start"), match.group("end"))


def minutes_to_text(minutes: int) -> str:
    suffix = "AM" if minutes < 12 * 60 else "PM"
    hour = (minutes // 60) % 12
    hour = 12 if hour == 0 else hour
    return f"{hour}:{minutes % 60:02d} {suffix}"


def remove_time_ranges(text: str) -> str:
    text = TIME_RANGE_RE.sub(" ", clean_text(text))
    text = re.sub(r"\(\s*\)", " ", text)
    return re.sub(r"\s+", " ", text).strip(" -–—:;")


def extract_teacher(text: str):
    compact = clean_text(text).strip(" -–—:;")
    match = re.search(
        r"([A-Za-z][A-Za-z.' ]{0,55}?\b(?:Sir|Ma['’]?am|Maam|Mam|Madam))\b",
        compact,
        flags=re.IGNORECASE,
    )
    return clean_text(match.group(1)) if match else None


def extract_subject_teacher(entry: str):
    text = remove_time_ranges(entry)
    if not looks_like_teacher_name(text):
        return None, None

    for subject, pattern in SUBJECT_PATTERNS:
        match = re.match(pattern, text, flags=re.IGNORECASE)
        if match:
            teacher = extract_teacher(text[match.end():])
            if teacher:
                return subject, teacher

    generic = re.match(r"^(?P<subject>[A-Za-z& ]{2,50}?)\s*[-:–—]\s*(?P<rest>.+)$", text)
    if generic:
        teacher = extract_teacher(generic.group("rest"))
        if teacher:
            return clean_text(generic.group("subject")).title(), teacher

    return None, None


def is_label_or_heading(value: str) -> bool:
    lower = clean_text(value).lower()
    if not lower:
        return True
    unwanted = (
        "batch name", "faculty name", "class", "weekly schedule",
        "sri chaitanya", "chandigarh, india", "class schedule",
        "permanent week off", "batch wise lecture count",
    )
    return any(term == lower or term in lower for term in unwanted)


def read_workbook_to_matrices(file_bytes: bytes, uploaded_name: str) -> dict[str, list[list]]:
    """
    Load XLSX/XLSM with openpyxl. XLS is supported with pandas + xlrd if installed.
    Merged cells are filled with their visible top-left value, which is important
    for vertical batch-name cells.
    """
    suffix = Path(uploaded_name).suffix.lower()

    if suffix == ".xls":
        try:
            workbook = pd.ExcelFile(BytesIO(file_bytes), engine="xlrd")
        except Exception as exc:
            raise ValueError(
                "This .xls file could not be read. Save it once as .xlsx, then upload it again."
            ) from exc

        result = {}
        for sheet_name in workbook.sheet_names:
            dataframe = workbook.parse(sheet_name, header=None, dtype=object)
            dataframe = dataframe.where(pd.notna(dataframe), None)
            result[sheet_name] = dataframe.values.tolist()
        return result

    if suffix not in {".xlsx", ".xlsm"}:
        raise ValueError("Upload an Excel workbook in .xlsx, .xlsm or .xls format.")

    try:
        workbook = load_workbook(
            BytesIO(file_bytes),
            data_only=True,
            read_only=False,
            keep_vba=False,
        )
    except Exception as exc:
        raise ValueError("The uploaded Excel file could not be opened.") from exc

    result = {}
    for worksheet in workbook.worksheets:
        max_row = worksheet.max_row
        max_col = worksheet.max_column
        matrix = [
            [worksheet.cell(row=row, column=col).value for col in range(1, max_col + 1)]
            for row in range(1, max_row + 1)
        ]

        # Replicate each merged cell's displayed value across the merged area.
        for merged_range in worksheet.merged_cells.ranges:
            top_value = worksheet.cell(merged_range.min_row, merged_range.min_col).value
            for row in range(merged_range.min_row, merged_range.max_row + 1):
                for col in range(merged_range.min_col, merged_range.max_col + 1):
                    matrix[row - 1][col - 1] = top_value

        result[worksheet.title] = matrix

    workbook.close()
    return result


def find_schedule_headers(matrix: list[list]):
    """
    Keep only the real header row. Some templates repeat 'TIMINGS/ DATE-DAY'
    in the row below the weekday labels; that second row is the date row, not
    a new schedule block.
    """
    headers = []
    for row_index, row in enumerate(matrix):
        for col_index, value in enumerate(row):
            text = clean_text(value).lower()
            if "timing" in text and ("date" in text or "day" in text):
                weekday_count = sum(
                    1
                    for candidate in row[col_index + 1:]
                    if get_day_name(candidate)
                )
                if weekday_count >= 2:
                    headers.append((row_index, col_index))
                break
    return headers


def find_day_columns(matrix: list[list], header_row: int, time_col: int):
    result = {}
    for col in range(time_col + 1, len(matrix[header_row])):
        day_name = get_day_name(cell_value(matrix, header_row, col))
        if day_name:
            result[col] = day_name
    return result


def derive_block_dates(matrix: list[list], header_row: int, day_columns: dict[int, str]):
    candidates = []
    for date_row in (header_row + 1, header_row + 2):
        for col, day_name in day_columns.items():
            parsed = parse_date_value(cell_value(matrix, date_row, col))
            if parsed:
                candidates.append((col, day_name, parsed))

    if not candidates:
        return {}, None

    anchor = next(
        (item for item in candidates if item[2].weekday() == DAY_INDEX[item[1]]),
        candidates[0],
    )
    _, anchor_day, anchor_date = anchor
    anchor_index = DAY_INDEX[anchor_day]
    return {
        col: anchor_date + timedelta(days=DAY_INDEX[day_name] - anchor_index)
        for col, day_name in day_columns.items()
    }, None


def find_batch_name(matrix: list[list], header_row: int, batch_col: int) -> str:
    current = clean_text(cell_value(matrix, header_row, batch_col))
    if current and not is_label_or_heading(current):
        return current

    for row in range(header_row - 1, max(-1, header_row - 11), -1):
        candidate = clean_text(cell_value(matrix, row, batch_col))
        if candidate and not is_label_or_heading(candidate):
            return candidate
    return ""


def is_faculty_schedule_block(matrix: list[list], header_row: int, batch_col: int) -> bool:
    return "faculty" in clean_text(cell_value(matrix, header_row, batch_col)).lower()


def extract_from_sheet(sheet_name: str, matrix: list[list]):
    sessions = []
    review_rows = []
    skipped_faculty_blocks = 0
    headers = find_schedule_headers(matrix)

    for header_number, (header_row, time_col) in enumerate(headers):
        block_end = headers[header_number + 1][0] if header_number + 1 < len(headers) else len(matrix)
        batch_col = max(0, time_col - 1)

        if is_faculty_schedule_block(matrix, header_row, batch_col):
            skipped_faculty_blocks += 1
            continue

        day_columns = find_day_columns(matrix, header_row, time_col)
        if not day_columns:
            continue

        date_map, _ = derive_block_dates(matrix, header_row, day_columns)
        if not date_map:
            review_rows.append([
                sheet_name, f"Row {header_row + 1}", "", "", "", "", "",
                "Could not read the date row for this schedule block.",
            ])
            continue

        batch_name = find_batch_name(matrix, header_row, batch_col)

        for row in range(header_row + 1, block_end):
            raw_slot = clean_text(cell_value(matrix, row, time_col))
            row_time = time_range_from_text(raw_slot)
            if not row_time:
                continue

            row_batch = clean_text(cell_value(matrix, row, batch_col))
            if row_batch and not is_label_or_heading(row_batch):
                batch_name = row_batch

            if not batch_name:
                review_rows.append([
                    sheet_name, f"{excel_column_letter(time_col + 1)}{row + 1}",
                    "", "", "", raw_slot, "",
                    "Batch name could not be detected for this timetable block.",
                ])
                continue

            for day_col, day_name in day_columns.items():
                entry = clean_text(cell_value(matrix, row, day_col))
                if not entry or is_skip_entry(entry):
                    continue

                source_cell = f"{excel_column_letter(day_col + 1)}{row + 1}"
                session_date = date_map[day_col]
                cell_time = time_range_from_text(entry)
                start_minutes, end_minutes = cell_time or row_time

                subject, teacher = extract_subject_teacher(entry)
                if not teacher:
                    review_rows.append([
                        sheet_name, source_cell, session_date, day_name.title(), batch_name,
                        raw_slot, entry,
                        "Teacher name could not be identified. Check this cell manually.",
                    ])
                    continue
                if not subject:
                    review_rows.append([
                        sheet_name, source_cell, session_date, day_name.title(), batch_name,
                        raw_slot, entry,
                        "Subject could not be identified. Check this cell manually.",
                    ])
                    continue

                sessions.append({
                    "date": session_date,
                    "day": day_name.title(),
                    "batch": batch_name,
                    "start_minutes": start_minutes,
                    "end_minutes": end_minutes,
                    "subject": subject,
                    "teacher": teacher,
                    "teacher_key": normalise_teacher(teacher),
                    "source_sheet": sheet_name,
                    "source_cell": source_cell,
                    "raw_entry": entry,
                })

    return sessions, review_rows, skipped_faculty_blocks


def deduplicate_sessions(sessions: list[dict]) -> list[dict]:
    unique = {}
    for item in sessions:
        key = (
            item["date"], item["batch"].casefold(),
            item["start_minutes"], item["end_minutes"],
            item["subject"].casefold(), item["teacher_key"],
        )
        if key not in unique:
            item["source_sheets"] = [item["source_sheet"]]
            item["source_cells"] = [item["source_cell"]]
            item["clashes"] = []
            unique[key] = item
        else:
            existing = unique[key]
            if item["source_sheet"] not in existing["source_sheets"]:
                existing["source_sheets"].append(item["source_sheet"])
            if item["source_cell"] not in existing["source_cells"]:
                existing["source_cells"].append(item["source_cell"])
    return list(unique.values())


def sort_sessions(sessions: list[dict]) -> list[dict]:
    return sorted(
        sessions,
        key=lambda item: (
            item["date"], item["start_minutes"],
            item["batch"].casefold(), item["teacher"].casefold(),
        ),
    )


def find_teacher_clashes(sessions: list[dict]) -> list[dict]:
    groups = defaultdict(list)
    for index, session in enumerate(sessions):
        groups[(session["date"], session["teacher_key"])].append(index)

    clash_pairs = []
    seen_pairs = set()

    for indexes in groups.values():
        indexes.sort(key=lambda idx: (sessions[idx]["start_minutes"], sessions[idx]["end_minutes"]))

        for position, left_index in enumerate(indexes):
            left = sessions[left_index]
            for right_index in indexes[position + 1:]:
                right = sessions[right_index]
                if right["start_minutes"] >= left["end_minutes"]:
                    break

                overlap_start = max(left["start_minutes"], right["start_minutes"])
                overlap_end = min(left["end_minutes"], right["end_minutes"])
                if overlap_end <= overlap_start:
                    continue

                pair = tuple(sorted((left_index, right_index)))
                if pair in seen_pairs:
                    continue
                seen_pairs.add(pair)

                left["clashes"].append(right_index)
                right["clashes"].append(left_index)
                clash_pairs.append({
                    "date": left["date"],
                    "day": left["day"],
                    "teacher": left["teacher"],
                    "batch_1": left["batch"],
                    "subject_1": left["subject"],
                    "start_1": left["start_minutes"],
                    "end_1": left["end_minutes"],
                    "batch_2": right["batch"],
                    "subject_2": right["subject"],
                    "start_2": right["start_minutes"],
                    "end_2": right["end_minutes"],
                    "overlap_minutes": overlap_end - overlap_start,
                    "sources": (
                        f"{', '.join(left['source_sheets'])} ({', '.join(left['source_cells'])})"
                        f" | {', '.join(right['source_sheets'])} ({', '.join(right['source_cells'])})"
                    ),
                })

    return clash_pairs


def extract_schedule(matrices: dict[str, list[list]], selected_sheets: list[str]):
    all_sessions = []
    all_review_rows = []
    skipped_faculty_blocks = 0

    for sheet_name in selected_sheets:
        sessions, review_rows, skipped = extract_from_sheet(sheet_name, matrices[sheet_name])
        all_sessions.extend(sessions)
        all_review_rows.extend(review_rows)
        skipped_faculty_blocks += skipped

    sessions = sort_sessions(deduplicate_sessions(all_sessions))
    for session in sessions:
        session["clashes"] = []
    clashes = find_teacher_clashes(sessions)

    return sessions, clashes, all_review_rows, skipped_faculty_blocks


def _base_formats(workbook):
    return {
        "title": workbook.add_format({
            "bold": True, "font_color": "#FFFFFF", "bg_color": "#7030A0",
            "align": "center", "valign": "vcenter", "border": 1,
        }),
        "header": workbook.add_format({
            "bold": True, "font_color": "#FFFFFF", "bg_color": "#7030A0",
            "align": "center", "valign": "vcenter", "border": 1, "text_wrap": True,
        }),
        "text": workbook.add_format({"border": 1, "valign": "vcenter"}),
        "center": workbook.add_format({"border": 1, "align": "center", "valign": "vcenter"}),
        "date": workbook.add_format({"border": 1, "num_format": "dd.mm.yyyy", "valign": "vcenter"}),
        "time": workbook.add_format({"border": 1, "num_format": "h:mm AM/PM", "align": "center", "valign": "vcenter"}),
        "clash_text": workbook.add_format({"border": 1, "bg_color": "#FFC7CE", "font_color": "#9C0006", "bold": True, "valign": "vcenter"}),
        "clash_center": workbook.add_format({"border": 1, "bg_color": "#FFC7CE", "font_color": "#9C0006", "bold": True, "align": "center", "valign": "vcenter"}),
        "clash_date": workbook.add_format({"border": 1, "bg_color": "#FFC7CE", "font_color": "#9C0006", "bold": True, "num_format": "dd.mm.yyyy", "valign": "vcenter"}),
        "clash_time": workbook.add_format({"border": 1, "bg_color": "#FFC7CE", "font_color": "#9C0006", "bold": True, "num_format": "h:mm AM/PM", "align": "center", "valign": "vcenter"}),
        "note": workbook.add_format({"border": 1, "text_wrap": True, "valign": "top"}),
    }


def _set_columns(worksheet, widths: list[int]):
    for index, width in enumerate(widths):
        worksheet.set_column(index, index, width)


def _write_schedule_sheet(worksheet, formats, sessions):
    headers = [
        "Date", "Day", "Batch", "Start Time", "End Time", "Subject",
        "Teacher Name", "Clash Status", "Clash With", "Source Sheet", "Source Cell",
    ]
    worksheet.write_row(0, 0, headers, formats["header"])
    worksheet.freeze_panes(1, 0)
    worksheet.autofilter(0, 0, max(1, len(sessions)), len(headers) - 1)
    _set_columns(worksheet, [13, 12, 28, 14, 14, 22, 24, 14, 44, 22, 14])

    for output_row, item in enumerate(sessions, start=1):
        is_clash = bool(item["clashes"])
        text_fmt = formats["clash_text"] if is_clash else formats["text"]
        center_fmt = formats["clash_center"] if is_clash else formats["center"]
        date_fmt = formats["clash_date"] if is_clash else formats["date"]
        time_fmt = formats["clash_time"] if is_clash else formats["time"]

        details = []
        for other_index in item["clashes"]:
            other = sessions[other_index]
            details.append(
                f"Row {other_index + 2}: {other['batch']} "
                f"({minutes_to_text(other['start_minutes'])}-{minutes_to_text(other['end_minutes'])})"
            )

        worksheet.write_datetime(output_row, 0, datetime.combine(item["date"], datetime.min.time()), date_fmt)
        worksheet.write(output_row, 1, item["day"], center_fmt)
        worksheet.write(output_row, 2, item["batch"], text_fmt)
        worksheet.write_datetime(output_row, 3, datetime(1899, 12, 30) + timedelta(minutes=item["start_minutes"]), time_fmt)
        worksheet.write_datetime(output_row, 4, datetime(1899, 12, 30) + timedelta(minutes=item["end_minutes"]), time_fmt)
        worksheet.write(output_row, 5, item["subject"], text_fmt)
        worksheet.write(output_row, 6, item["teacher"], text_fmt)
        worksheet.write(output_row, 7, "CLASH" if is_clash else "OK", center_fmt)
        worksheet.write(output_row, 8, " | ".join(details), text_fmt)
        worksheet.write(output_row, 9, ", ".join(item["source_sheets"]), text_fmt)
        worksheet.write(output_row, 10, ", ".join(item["source_cells"]), center_fmt)


def _write_clashes_sheet(worksheet, formats, clash_pairs):
    headers = [
        "Date", "Day", "Teacher Name",
        "Batch 1", "Subject 1", "Start Time 1", "End Time 1",
        "Batch 2", "Subject 2", "Start Time 2", "End Time 2",
        "Overlap (Minutes)", "Source Locations",
    ]
    worksheet.write_row(0, 0, headers, formats["header"])
    worksheet.freeze_panes(1, 0)
    worksheet.autofilter(0, 0, max(1, len(clash_pairs)), len(headers) - 1)
    _set_columns(worksheet, [13, 12, 24, 28, 22, 14, 14, 28, 22, 14, 14, 18, 44])

    for row, item in enumerate(clash_pairs, start=1):
        worksheet.write_datetime(row, 0, datetime.combine(item["date"], datetime.min.time()), formats["date"])
        worksheet.write(row, 1, item["day"], formats["center"])
        worksheet.write(row, 2, item["teacher"], formats["text"])
        worksheet.write(row, 3, item["batch_1"], formats["text"])
        worksheet.write(row, 4, item["subject_1"], formats["text"])
        worksheet.write_datetime(row, 5, datetime(1899, 12, 30) + timedelta(minutes=item["start_1"]), formats["time"])
        worksheet.write_datetime(row, 6, datetime(1899, 12, 30) + timedelta(minutes=item["end_1"]), formats["time"])
        worksheet.write(row, 7, item["batch_2"], formats["text"])
        worksheet.write(row, 8, item["subject_2"], formats["text"])
        worksheet.write_datetime(row, 9, datetime(1899, 12, 30) + timedelta(minutes=item["start_2"]), formats["time"])
        worksheet.write_datetime(row, 10, datetime(1899, 12, 30) + timedelta(minutes=item["end_2"]), formats["time"])
        worksheet.write(row, 11, item["overlap_minutes"], formats["center"])
        worksheet.write(row, 12, item["sources"], formats["note"])


def _write_review_sheet(worksheet, formats, review_rows):
    headers = ["Source Sheet", "Source Cell", "Date", "Day", "Batch", "Slot Time", "Original Entry", "Reason"]
    worksheet.write_row(0, 0, headers, formats["header"])
    worksheet.freeze_panes(1, 0)
    worksheet.autofilter(0, 0, max(1, len(review_rows)), len(headers) - 1)
    _set_columns(worksheet, [22, 14, 13, 12, 28, 16, 46, 56])

    for row, item in enumerate(review_rows, start=1):
        worksheet.write(row, 0, item[0], formats["text"])
        worksheet.write(row, 1, item[1], formats["center"])
        if isinstance(item[2], date):
            worksheet.write_datetime(row, 2, datetime.combine(item[2], datetime.min.time()), formats["date"])
        else:
            worksheet.write(row, 2, item[2], formats["text"])
        worksheet.write(row, 3, item[3], formats["center"])
        worksheet.write(row, 4, item[4], formats["text"])
        worksheet.write(row, 5, item[5], formats["center"])
        worksheet.write(row, 6, item[6], formats["note"])
        worksheet.write(row, 7, item[7], formats["note"])


def _write_summary_sheet(worksheet, formats, source_excel_base, selected_sheets, sessions, clashes, reviews, skipped):
    rows = [
        ["Source Excel Name", source_excel_base],
        ["Input tabs selected", ", ".join(selected_sheets)],
        ["Total extracted sessions", len(sessions)],
        ["Teacher clash pairs", len(clashes)],
        ["Rows requiring review", len(reviews)],
        ["Faculty/team blocks skipped", skipped],
        ["Output file type", "Macro-enabled Excel workbook (.xlsm)"],
        ["Last/source sheet", "Schedule - do not rename it; the VBA reads this sheet."],
        ["Create schedules + individual PDFs", "Open the XLSM, click Enable Content, press Alt + F8 and run Create_All_Schedules."],
        ["Teacher PDF folder", f"{source_excel_base}_teacher schedule - one PDF per teacher"],
        ["Batch PDF folder", f"{source_excel_base}_Batchwise schedule - one PDF per batch"],
        ["Generated on", datetime.now().strftime("%d.%m.%Y %I:%M %p")],
    ]
    worksheet.write_row(0, 0, ["Item", "Value"], formats["header"])
    _set_columns(worksheet, [30, 95])
    worksheet.set_row(0, 24)
    for row, values in enumerate(rows, start=1):
        worksheet.write(row, 0, values[0], formats["text"])
        worksheet.write(row, 1, values[1], formats["note"])


def build_macro_workbook(
    sessions: list[dict],
    clash_pairs: list[dict],
    review_rows: list[list],
    selected_sheets: list[str],
    skipped_blocks: int,
    source_excel_base: str,
    vba_project_path: str | Path,
) -> bytes:
    """
    Build an XLSM using the user's real vbaProject.bin. It contains no source
    workbook data beyond the extracted combined schedule.
    """
    output = BytesIO()
    workbook = xlsxwriter.Workbook(output, {"in_memory": True})
    workbook.set_properties({
        "title": "Weekly Schedule Extractor",
        "subject": "Combined schedule with teacher clash checking",
        "author": "Sri Chaitanya Chandigarh",
    })
    workbook.add_vba_project(str(vba_project_path))
    formats = _base_formats(workbook)

    # Order matters. Schedule must remain the final sheet for the embedded VBA.
    summary_ws = workbook.add_worksheet("Summary")
    clashes_ws = workbook.add_worksheet("Clashes Only")
    review_ws = workbook.add_worksheet("Review Needed")
    schedule_ws = workbook.add_worksheet("Schedule")

    _write_summary_sheet(
        summary_ws, formats, source_excel_base, selected_sheets,
        sessions, clash_pairs, review_rows, skipped_blocks,
    )
    _write_clashes_sheet(clashes_ws, formats, clash_pairs)
    _write_review_sheet(review_ws, formats, review_rows)
    _write_schedule_sheet(schedule_ws, formats, sessions)

    workbook.close()
    return output.getvalue()


def suggested_batch_tabs(sheet_names: list[str]) -> list[str]:
    selected = [
        sheet for sheet in sheet_names
        if not any(term in sheet.lower() for term in ("faculty", "teacher", "team"))
    ]
    return selected or sheet_names
